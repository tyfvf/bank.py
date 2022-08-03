from tkinter import messagebox
from openpyxl import load_workbook

class Bank:

    def __init__(self, username: str, window, password = '', money=0):
        self.__window = window
        self.__username = username
        self.__password = password
        self.__money = money


    def sign_up(self):
        wk = load_workbook('users.xlsx')
        sheet = wk.active
        
        for cell in sheet['A']:
            row = cell.row + 1

            if cell.value == self.__username:
                messagebox.showerror('Used username', 'This username already exists, please choose another one')
                return
            
            if sheet[f'A{row}'].value == None:
                sheet[f'A{row}'].value = self.__username

            if sheet[f'B{row}'].value == None:
                sheet[f'B{row}'].value = self.__password

            if sheet[f'C{row}'].value == None:
                sheet[f'C{row}'].value = self.__money
            

        wk.save('users.xlsx')
        messagebox.showinfo('Signed up!', 'You signed up with sucess!')
        self.__window.destroy()


    def log_in(self):
        wk = load_workbook('users.xlsx')
        sheet = wk.active

        for cell in sheet['A']:
            if cell.value == self.__username and sheet[f'B{cell.row}'].value == self.__password:
                return True
            
        return False


    def deposit(self):
        try:

            number = float(self.__money)
            if number <= 0:
                return messagebox.showerror('ERROR', 'Please enter a valid number')
            else:
                wk = load_workbook('users.xlsx')
                sheet = wk.active

                for cell in sheet['A']:
                    if cell.value == self.__username:
                        sheet[f'C{cell.row}'].value += number
                
                wk.save('users.xlsx')
                messagebox.showinfo('Sucess!', 'Value deposited with sucess!')
                self.__window.destroy()

        except ValueError:
            return messagebox.showerror('ERROR', 'Please enter a valid number (if the number you are putting in is a decimal, please instead of using comma use a dot, thanks)')


    def transfer(self, to):
        try:

            number = float(self.__money)
            if number <= 0:
                return messagebox.showerror('ERROR', 'Please enter a valid number')
            else:
                wk = load_workbook('users.xlsx')
                sheet = wk.active

                exist_to = False
                enough_money = False

                for cell in sheet['A']:
                    if cell.value == to:
                        exist_to = True
                    elif cell.value == self.__username:
                        if sheet[f'C{cell.row}'].value >= number:
                            enough_money = True

                if exist_to and enough_money:
                    for cell in sheet['A']:
                        if cell.value == self.__username:
                            sheet[f'C{cell.row}'].value -= number
                        elif cell.value == to:
                            sheet[f'C{cell.row}'].value += number
                elif exist_to and enough_money == False:
                    return messagebox.showerror('Error!', 'You do not have enough money to make this transaction, please enter a lower number')
                elif enough_money and exist_to == False:
                    return messagebox.showerror('Error!', 'This username does not exist, please enter a valid username')
                else:
                    return messagebox.showerror('Error!', 'You do not have enough money and this username does not exist')

                
                wk.save('users.xlsx')
                messagebox.showinfo('Sucess!', 'Value transfered with sucess!')
                self.__window.destroy()

        except ValueError:
            return messagebox.showerror('ERROR', 'Please enter a valid number (if the number you are putting in is a decimal, please instead of using comma use a dot, thanks)')


    def withdraw(self):
        try:

            number = float(self.__money)
            if number <= 0:
                return messagebox.showerror('ERROR', 'Please enter a valid number')
            else:
                wk = load_workbook('users.xlsx')
                sheet = wk.active

                for cell in sheet['A']:
                    if cell.value == self.__username:
                        if number > sheet[f'C{cell.row}'].value:
                            return messagebox.showerror('Error!', 'Can not withdraw more than your account have! Please enter a lower number')
                        sheet[f'C{cell.row}'].value -= number
                
                wk.save('users.xlsx')
                messagebox.showinfo('Sucess!', 'Value withdrew with sucess!')
                self.__window.destroy()

        except ValueError:
            return messagebox.showerror('ERROR', 'Please enter a valid number (if the number you are putting in is a decimal, please instead of using comma use a dot, thanks)')


    def statment(self):
        wk = load_workbook('users.xlsx')
        sheet = wk.active

        for cell in sheet['A']:
            if cell.value == self.__username:
                return sheet[f'C{cell.row}'].value
        